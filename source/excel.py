from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def get_excel(excursion_destinations) -> None:
    """Creates excel spreadsheet from list of destination objects.

    Args:
        excursion_destinations (list): List containing destination objects.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Destinations"

    ws.append(["Name", "Description", "Contact"])

    counter = 2
    for dest in excursion_destinations:
        ws[f'A{counter}'].hyperlink = dest.url
        ws[f'A{counter}'].value = dest.name
        ws[f'A{counter}'].style = "Hyperlink"
        ws[f'A{counter}'].alignment = Alignment(vertical='center')
        ws[f'B{counter}'].value = dest.description
        ws[f'B{counter}'].alignment = Alignment(wrap_text=True, vertical='center')
        ws[f'C{counter}'].value = str.join("\n", dest.contact_info)
        ws[f'C{counter}'].alignment = Alignment(wrap_text=True, vertical='center')
        counter += 1

    ft = Font(bold=True)
    for row in ws["A1:C1"]:
        for cell in row:
            cell.font = ft

    max_name, max_text, max_contact = get_column_widths(excursion_destinations)
    ws.column_dimensions["A"].width = max_name
    ws.column_dimensions["B"].width = max_text
    ws.column_dimensions["C"].width = max_contact

    #TODO rewrite to use other path
    wb.save('./output/destinations.xlsx')

def get_column_widths(excursion_destinations) -> tuple:
    """Returns suggested column width for 3 rows present in excel file.

    Args:
        excursion_destinations (list): List containing destination objects.

    Returns:
        tuple: Tuple containing suggested column widths.
    """
    max_name = max([len(i.name) for i in excursion_destinations])
    max_contact = max([len(max(i.contact_info, key=len)) for i in excursion_destinations])
    max_text = max(max_name, max_contact)
    return max_name + 2, max_text +2, max_contact + 2
